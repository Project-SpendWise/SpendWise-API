import pandas as pd
import sys
import zipfile
from xml.etree import ElementTree as ET

def convert_xlsx_to_csv_manual(xlsx_file, csv_file):
    """
    Manually extract data from xlsx file by reading the XML directly
    This bypasses all styling issues
    """
    import openpyxl
    from openpyxl import load_workbook
    
    # Patch the Fill class to ignore style errors
    import openpyxl.styles.fills
    original_fill_init = openpyxl.styles.fills.Fill.__init__
    
    def patched_fill_init(self, *args, **kwargs):
        try:
            original_fill_init(self, *args, **kwargs)
        except:
            # If styling fails, just create empty Fill
            self.patternType = None
            self.fgColor = None
            self.bgColor = None
    
    openpyxl.styles.fills.Fill.__init__ = patched_fill_init
    
    # Now try to load
    wb = load_workbook(xlsx_file, data_only=True)
    ws = wb.active
    
    # Extract all data
    data = []
    for row in ws.iter_rows(values_only=True):
        data.append(row)
    
    # Create DataFrame
    if data:
        df = pd.DataFrame(data[1:], columns=data[0])
        df.to_csv(csv_file, index=False, encoding='utf-8-sig')
        return df
    return None

def convert_xlsx_to_csv(xlsx_file, csv_file=None):
    """
    Convert Excel file to CSV format
    
    Args:
        xlsx_file: Path to the input Excel file
        csv_file: Path to the output CSV file (optional)
    """
    # If no output file specified, create one based on input filename
    if csv_file is None:
        csv_file = xlsx_file.rsplit('.', 1)[0] + '.csv'
    
    try:
        # Try pandas first (fastest)
        df = pd.read_excel(xlsx_file, engine='openpyxl')
        df.to_csv(csv_file, index=False, encoding='utf-8-sig')
        
        print(f"✓ Successfully converted '{xlsx_file}' to '{csv_file}'")
        print(f"  Rows: {len(df)}")
        print(f"  Columns: {len(df.columns)}")
        print(f"  Column names: {', '.join(df.columns.tolist())}")
        return
        
    except Exception as e:
        print(f"Standard method failed: {str(e)}")
        print("Trying patched method...")
    
    try:
        # Try with patched openpyxl
        df = convert_xlsx_to_csv_manual(xlsx_file, csv_file)
        
        if df is not None:
            print(f"✓ Successfully converted '{xlsx_file}' to '{csv_file}'")
            print(f"  Rows: {len(df)}")
            print(f"  Columns: {len(df.columns)}")
            print(f"  Column names: {', '.join(df.columns.tolist())}")
            return
            
    except Exception as e:
        print(f"Patched method failed: {str(e)}")
        print("Trying LibreOffice method...")
    
    try:
        # Last resort: try using pyexcel-ods
        import pyexcel as pe
        sheet = pe.get_sheet(file_name=xlsx_file)
        sheet.save_as(csv_file)
        
        print(f"✓ Successfully converted '{xlsx_file}' to '{csv_file}'")
        return
        
    except ImportError:
        print("✗ All methods failed.")
        print("\nPlease try installing additional libraries:")
        print("  pip install pyexcel pyexcel-xlsx")
        print("\nOr use LibreOffice command line:")
        print(f"  soffice --headless --convert-to csv '{xlsx_file}'")
    except Exception as e:
        print(f"✗ Final error: {str(e)}")

if __name__ == "__main__":
    # Check if filename was provided
    if len(sys.argv) < 2:
        print("Usage: python script.py <input_file.xlsx> [output_file.csv]")
        print("Example: python script.py bank_transactions.xlsx")
        print("Example: python script.py bank_transactions.xlsx output.csv")
        sys.exit(1)
    
    # Get input file
    input_file = sys.argv[1]
    
    # Get output file if provided
    output_file = sys.argv[2] if len(sys.argv) > 2 else None
    
    # Convert the file
    convert_xlsx_to_csv(input_file, output_file)